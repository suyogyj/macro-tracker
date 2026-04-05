#!/usr/bin/env python3
"""Convert INDB.xlsx (Nutrient Data sheet) to data/indb-foods.json for the macro tracker."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "INDB.xlsx"
OUT_PATH = ROOT / "data" / "indb-foods.json"

# Column indices (0-based) from header row
COL_CODE = 0
COL_NAME = 1
COL_ENERGY_KCAL = 4
COL_CARB = 5
COL_PROTEIN = 6
COL_FAT = 7
COL_SERVINGS_UNIT = 42
COL_UNIT_KCAL = 44


def slug_external_key(food_code: str | None, name: str) -> str:
    code = (food_code or "").strip().lower()
    if code:
        return f"indb:{code}"
    safe = re.sub(r"[^a-z0-9]+", "-", (name or "food").lower()).strip("-")
    return f"indb:name:{safe[:80]}"


def parse_grams_from_label(s: str | None) -> float | None:
    if not s or not isinstance(s, str):
        return None
    m = re.search(r"(\d+(?:\.\d+)?)\s*g(?:ram)?s?\b", s, re.I)
    if m:
        return float(m.group(1))
    return None


def serving_grams(row: tuple) -> tuple[float | None, str | None]:
    """Infer reference grams for unit_serving macros; return (grams, label)."""
    label = row[COL_SERVINGS_UNIT]
    label_str = str(label).strip() if label is not None else ""
    energy_100 = float(row[COL_ENERGY_KCAL] or 0)
    unit_kcal = row[COL_UNIT_KCAL]
    unit_kcal_f = float(unit_kcal) if unit_kcal is not None else 0.0

    g = parse_grams_from_label(label_str)
    if g and g > 0:
        return g, label_str

    if energy_100 > 0 and unit_kcal_f > 0:
        g = 100.0 * unit_kcal_f / energy_100
        if 1 <= g <= 5000:
            return g, label_str

    return None, label_str if label_str else None


def round_n(x: float, places: int = 2) -> float:
    if x is None:
        return 0.0
    return round(float(x), places)


def row_to_food(row: tuple) -> dict | None:
    name = row[COL_NAME]
    if not name or not str(name).strip():
        return None

    code = row[COL_CODE]
    external_key = slug_external_key(str(code) if code else None, str(name))

    cal = round_n(row[COL_ENERGY_KCAL] or 0, 1)
    carb = round_n(row[COL_CARB] or 0, 2)
    protein = round_n(row[COL_PROTEIN] or 0, 2)
    fat = round_n(row[COL_FAT] or 0, 2)

    serving_options = [{"label": "100g", "weight": 100}]
    grams, su_label = serving_grams(row)
    if grams and su_label:
        w = max(1, int(round(grams)))
        label = f"1 × {su_label}" if not su_label.lower().startswith("1") else su_label
        if f"({w}g)" not in label.lower():
            label = f"{label} (~{w}g)"
        serving_options.append({"label": label[:120], "weight": float(w)})

    return {
        "name": str(name).strip(),
        "protein": protein,
        "carbs": carb,
        "fats": fat,
        "calories": cal,
        "category": "indb",
        "servingOptions": serving_options,
        "isDefault": True,
        "source": "indb",
        "externalKey": external_key,
    }


def main() -> None:
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if not XLSX_PATH.exists():
        print(f"Missing {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, values_only=True)

    foods: list[dict] = []
    seen: set[str] = set()
    for row in rows:
        food = row_to_food(row)
        if not food:
            continue
        key = food["externalKey"]
        if key in seen:
            continue
        seen.add(key)
        foods.append(food)

    wb.close()

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(foods, indent=2), encoding="utf-8")
    print(f"Wrote {len(foods)} foods to {OUT_PATH}")


if __name__ == "__main__":
    main()
